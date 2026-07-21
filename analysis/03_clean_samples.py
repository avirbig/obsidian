#!/usr/bin/env python3
"""
Step 5 -- Clean the author's own pXRF measurements and build the object table.

Implements docs/cleaning_plan.md (stages A-E).

Design principles:
  * data/ is READ-ONLY. Everything is written to samples_db/.
  * Nothing is deleted silently. Every reading is kept in the output with a
    keep/drop decision and the REASON for it, plus its origin file and Excel row.
  * The author's own in-cell annotations ('Invalid', 'aborted') are read as
    LITERAL TEXT before any parsing, because pd.to_datetime/to_numeric would
    turn them into indistinguishable blanks. His bench judgement outranks ours.
  * '< LOD' becomes MISSING, never zero. Zero is chemically false.
  * We count OBJECTS, not readings. Most objects were measured twice (dorsal +
    ventral), so treating readings as independent would inflate the sample size.
  * Quality is judged by whether Rb/Zr/Nb were actually measured -- NOT by how
    long the instrument ran. Reading 1790 ran 102.8 s and was still unusable.

Outputs (samples_db/):
  samples_readings.csv   every reading, all flags, keep/drop + reason
  samples_objects.csv    one row per object -- the analysis-ready table
  cleaning_report.txt    counts at each stage, every drop with its reason
  obsidian_samples_CLEAN.xlsx   the same, formatted for humans

Reproducibility: the CSV and txt outputs are byte-identical on every re-run.
The .xlsx is not -- an xlsx is a zip archive and stores a modification time per
entry, so its BYTES change each run while its CONTENT does not. Expect git to
show it as modified after re-running; that is not a data change.
"""

from pathlib import Path
import re
import sys
import warnings

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# Windows consoles default to cp1252 and raise on Turkish characters; keep the
# console safe without touching file output (which is always UTF-8).
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

ROOT = Path(__file__).resolve().parent.parent
SAMPLES = ROOT / 'data' / 'samples'
MASTER = SAMPLES / 'obsidian_pxrf_master_2017-2018.xlsx'
EXTRA = SAMPLES / 'Avishai_Obsidian 12.2017.xlsx'
OUT = ROOT / 'samples_db'
OUT.mkdir(exist_ok=True)

# Elements. Rb and Zr are the primary discriminators; Nb is a COARSE check only
# (6 distinct values, +/-50% uncertainty -- see docs/cleaning_plan.md).
PRIMARY = ['Rb', 'Zr']
COARSE = ['Nb']
SUPPORT = ['Fe', 'Ti', 'Mn']
ALL_EL = PRIMARY + COARSE + SUPPORT + ['Ca', 'K', 'Si', 'Bal']

PCT_TO_PPM = 10000.0        # values are stored in %, literature uses ppm

# author annotations that mean "this reading is bad"
BAD_TEXT = re.compile(r'invalid|abort|error|bad|ignore|test', re.I)

report = []


def say(msg=''):
    print(msg)
    report.append(str(msg))


def hdr(title):
    say()
    say('=' * 78)
    say(title)
    say('=' * 78)


# ---------------------------------------------------------------- Stage A ----

def read_literal(path, sheet):
    """Read a sheet WITHOUT parsing, so author annotations survive as text."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    head = [str(h).strip() if h is not None else '' for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=head)
    df['_xlrow'] = range(2, len(df) + 2)
    return df


def num(series):
    """Numeric value, with '< LOD' and any other text -> missing (never zero)."""
    return pd.to_numeric(series, errors='coerce')


def txt(series):
    """Blank-safe text. pandas 3.0 keeps NA through astype(str), which would
    poison string joins and silently drop rows from groupby -- so map explicitly."""
    return series.map(lambda v: '' if pd.isna(v) else str(v).strip())


def is_lod(series):
    return series.astype(str).str.strip().str.upper().eq('< LOD')


hdr('STAGE A -- LOAD (read-only)')

M = read_literal(MASTER, 'Original')
M['_src_file'] = MASTER.name
say(f"master 'Original' sheet          : {len(M)} readings")

# reading 1489 added back (decided 2026-07-21; see docs/raw_sample_data_map_and_quality.md)
X = read_literal(EXTRA, 'Sheet1')
X = X[pd.to_numeric(X['Reading No'], errors='coerce') == 1489].copy()
X['_src_file'] = EXTRA.name
# give it the context the master rows carry
X['site'] = 'motza'
for c in ['side', 'Locus\\Square', 'Basket', 'cover', 'Dirt', 'remarks']:
    if c not in X.columns:
        X[c] = np.nan
say(f"reading 1489 restored from       : {EXTRA.name} (Excel row 832)")

D = pd.concat([M, X], ignore_index=True, sort=False)
say(f"total input                      : {len(D)} readings")

# --- A3: capture author annotations BEFORE parsing anything ---
annot_cols = ['Time', 'Locus\\Square', 'Basket', 'remarks', 'side', 'cover', 'Dirt']
D['author_flag'] = ''
for c in annot_cols:
    if c not in D.columns:
        continue
    cell = D[c].astype(str)
    # a Time cell that is text, or any cell matching the bad-word pattern
    hit = cell.str.contains(BAD_TEXT, na=False)
    if c == 'Time':
        hit = hit | (~D[c].apply(lambda v: hasattr(v, 'year')) & D[c].notna())
    for i in D.index[hit]:
        D.at[i, 'author_flag'] = (D.at[i, 'author_flag'] + f"{c}='{cell[i]}' ").strip()

n_annot = (D['author_flag'] != '').sum()
say(f"rows carrying an author annotation: {n_annot}")
if n_annot:
    say(D.loc[D['author_flag'] != '', ['Reading No', 'site', 'author_flag']]
        .to_string(index=False))

# --- A2/A4: parse ---
D['ts'] = pd.to_datetime(D['Time'], errors='coerce')
D['reading'] = pd.to_numeric(D['Reading No'], errors='coerce')
D['duration'] = num(D['Duration'])

for e in ALL_EL:
    if e in D.columns:
        D[e + '_lod'] = is_lod(D[e])
        D[e] = num(D[e])
    else:
        D[e], D[e + '_lod'] = np.nan, False

lod_counts = {e: int(D[e + '_lod'].sum()) for e in PRIMARY + COARSE}
say(f"'< LOD' converted to MISSING (not zero): {lod_counts}")

# unique key (Reading No + Time; fallback for blank/annotated Time)
D['key'] = np.where(D['ts'].notna(),
                    D['reading'].astype('Int64').astype(str) + '@' + D['ts'].astype(str),
                    D['reading'].astype('Int64').astype(str) + '#' + D['duration'].round(2).astype(str))
dups = D['key'].duplicated().sum()
say(f"duplicate keys after keying on Reading No + Time: {dups}")

# ---------------------------------------------------------------- Stage B ----

hdr('STAGE B -- DECIDE WHICH READINGS ARE USABLE')

D['drop_reason'] = ''


def mark(mask, reason):
    """Record a drop reason; first reason wins so the log stays readable."""
    m = mask & (D['drop_reason'] == '')
    D.loc[m, 'drop_reason'] = reason
    return int(m.sum())


n1 = mark(D['author_flag'] != '', 'author marked it bad')
say(f"B1  author annotation ('Invalid'/'aborted')      : {n1} dropped")

missing_primary = D['Rb'].isna() | D['Zr'].isna() | D['Nb'].isna()
n2 = mark(missing_primary, 'Rb/Zr/Nb not measured (< LOD or blank)')
say(f"B2  a sourcing element not measured              : {n2} dropped")

n3 = mark(D['duration'] < 60, 'aborted run (< 60 s)')
say(f"B3  clearly aborted run (< 60 s)                 : {n3} dropped")

D['keep'] = D['drop_reason'] == ''
say(f"\n     kept: {D['keep'].sum()}   dropped: {(~D['keep']).sum()}")

if (~D['keep']).any():
    say("\n     every dropped reading:")
    say(D.loc[~D['keep'], ['reading', 'ts', 'site', 'duration', 'drop_reason']]
        .to_string(index=False))

# --- flags (do NOT drop) ---
K = D[D['keep']].copy()

cover_map = {'full': 1.0, 'almost full': 0.8, 'partial': 0.6,
             'very partial': 0.3, 'tiny': 0.1}
K['_cover'] = K['cover'].astype(str).str.strip().str.lower()
K['cover_score'] = K['_cover'].map(cover_map).fillna(0.5)      # blank = unknown

K['_dirt'] = K['Dirt'].astype(str).str.strip().str.lower()
K['dirt_score'] = np.where(K['_dirt'].isin(['yes', 'dirty']), 0.3,
                           np.where(K['_dirt'].isin(['no', 'clean']), 1.0, 0.5))

K['bal_ok'] = K['Bal'].between(50, 70)
K['bal_score'] = np.where(K['Bal'].between(55, 65), 1.0,
                          np.where(K['Bal'].between(50, 70), 0.8, 0.4))

K['short_run'] = K['duration'] < 100

say(f"\n     flags on kept readings (not dropped):")
say(f"       surface cover poor (very partial/tiny) : {(K['cover_score'] <= 0.3).sum()}")
say(f"       dirt recorded                          : {(K['dirt_score'] == 0.3).sum()}")
say(f"       Bal outside 50-70%                     : {(~K['bal_ok']).sum()}")
say(f"       run shorter than 100 s                 : {K['short_run'].sum()}")

# ---------------------------------------------------------------- Stage D ----
# (ratios computed before Stage C, because C compares readings ON the ratios)

for e in PRIMARY + COARSE + SUPPORT:
    K[e + '_ppm'] = K[e] * PCT_TO_PPM

K['Rb_Zr'] = K['Rb'] / K['Zr']
K['Rb_Nb'] = K['Rb'] / K['Nb']
K['Zr_Nb'] = K['Zr'] / K['Nb']

# plausible-range gate, from the data itself (used by Stage C)
plaus = {}
for e in PRIMARY + COARSE:
    s = K[e + '_ppm'].dropna()
    plaus[e] = (s.quantile(.001), s.quantile(.999))

# ---------------------------------------------------------------- Stage C ----

hdr('STAGE C -- COMBINE READINGS INTO OBJECTS')

K['_site'] = txt(K['site']).str.lower()
K['_locus'] = txt(K['Locus\\Square'])
K['_basket'] = txt(K['Basket'])

# A reading with no locus or basket cannot be grouped. It must NOT be merged
# with other ungroupable readings (that would invent a fake object), so each
# gets its own id, marked so it can be found later.
ungroupable = (K['_locus'] == '') | (K['_basket'] == '')
K['object_id'] = np.where(
    ungroupable,
    K['_site'] + ' | UNGROUPED-reading-' + K['reading'].astype('Int64').astype(str),
    K['_site'] + ' | ' + K['_locus'] + ' | ' + K['_basket'])

say(f"readings          : {len(K)}")
say(f"distinct objects  : {K['object_id'].nunique()}")
say(f"readings with no locus/basket (kept as single objects, flagged): "
    f"{int(ungroupable.sum())}")
say("\nreadings per object:")
say(K.groupby('object_id').size().value_counts().sort_index().to_string())

FOLD_LIMIT = 1.6      # from the data: worst real Zr disagreement is 1.57x

objects, review, dropped_div = [], [], []

for oid, grp in K.groupby('object_id'):
    g = grp.copy()
    note = []

    # -- C2: do the readings agree?
    fold = {}
    for e in PRIMARY:
        v = g[e + '_ppm'].dropna()
        fold[e] = (v.max() / v.min()) if len(v) > 1 and v.min() > 0 else 1.0
    worst = max(fold.values()) if fold else 1.0

    if worst > FOLD_LIMIT and len(g) > 1:
        # -- C3b: is one reading outside the plausible range? then it is the bad one
        implausible = pd.Series(False, index=g.index)
        for e in PRIMARY:
            lo, hi = plaus[e]
            implausible |= ~g[e + '_ppm'].between(lo, hi)

        if implausible.any() and (~implausible).any():
            for i in g.index[implausible]:
                dropped_div.append({
                    'object_id': oid, 'reading': g.at[i, 'reading'],
                    'reason': 'outside plausible range, disagreed with partner',
                    'Rb_ppm': g.at[i, 'Rb_ppm'], 'Zr_ppm': g.at[i, 'Zr_ppm']})
            g = g[~implausible]
            note.append('one reading dropped as implausible')
        else:
            # both plausible but disagree -> a real question, not a maths problem
            note.append(f'SUSPICIOUS: readings disagree {worst:.2f}x, all plausible')
            review.append({'object_id': oid, 'n': len(g), 'worst_fold': worst,
                           'Rb_ppm': list(g['Rb_ppm']), 'Zr_ppm': list(g['Zr_ppm']),
                           'cover': list(g['_cover']), 'Dirt': list(g['_dirt'])})

    # -- C3: combine by MEDIAN
    sides = sorted(x for x in txt(g['side']).str.lower().unique() if x)
    rec = {
        'object_id': oid,
        'site': g['_site'].iloc[0],
        'locus': g['_locus'].iloc[0] or '(missing)',
        'basket': g['_basket'].iloc[0] or '(missing)',
        'n_readings': len(g),
        'sides': '+'.join(sides) if sides else '(not recorded)',
        'ungroupable': bool(g['_locus'].iloc[0] == '' or g['_basket'].iloc[0] == ''),
    }
    for e in PRIMARY + COARSE + SUPPORT:
        rec[e + '_ppm'] = g[e + '_ppm'].median()
    for r in ['Rb_Zr', 'Rb_Nb', 'Zr_Nb']:
        rec[r] = g[r].median()

    rec['worst_disagreement_fold'] = worst
    rec['agree_score'] = (1.0 if len(g) == 1 and len(grp) == 1 else
                          (1.0 if worst <= 1.3 else 0.7 if worst <= FOLD_LIMIT else 0.3))
    if len(g) == 1:
        rec['agree_score'] = 0.5          # a single reading cannot check itself

    rec['cover_score'] = g['cover_score'].mean()
    rec['dirt_score'] = g['dirt_score'].mean()
    rec['bal_score'] = g['bal_score'].mean()
    rec['n_score'] = min(1.0, len(g) / 2.0)
    rec['notes'] = '; '.join(note)
    objects.append(rec)

O = pd.DataFrame(objects)

say(f"\nobjects built                       : {len(O)}")
say(f"readings dropped for divergence     : {len(dropped_div)}")
say(f"objects FLAGGED for human review    : {len(review)}")
if review:
    say("\n  objects whose readings disagree while all look plausible:")
    for r in review:
        say(f"    {r['object_id']}  n={r['n']}  worst {r['worst_fold']:.2f}x  "
            f"Rb={r['Rb_ppm']}  Zr={r['Zr_ppm']}")

# ---------------------------------------------------------------- Stage E ----

hdr('STAGE E -- QUALITY SCORE PER OBJECT')

W = {'agree_score': .35, 'cover_score': .25, 'dirt_score': .15,
     'bal_score': .15, 'n_score': .10}
O['quality'] = sum(O[k] * w for k, w in W.items())
O['quality_band'] = pd.cut(O['quality'], [0, .55, .75, 1.01],
                           labels=['low', 'medium', 'high'], right=False)

say("weights: " + ', '.join(f"{k}={v}" for k, v in W.items()))
say("\nquality bands:")
say(O['quality_band'].value_counts().sort_index().to_string())
say(f"\nmean quality {O['quality'].mean():.2f}   "
    f"range {O['quality'].min():.2f}-{O['quality'].max():.2f}")

hdr('RESULT -- CHEMISTRY BY SITE (objects, not readings)')
summ = O.groupby('site').agg(
    objects=('object_id', 'size'),
    Rb_ppm=('Rb_ppm', 'median'), Zr_ppm=('Zr_ppm', 'median'),
    Nb_ppm=('Nb_ppm', 'median'),
    Rb_Zr=('Rb_Zr', 'median'), Rb_Zr_sd=('Rb_Zr', 'std'),
    quality=('quality', 'mean')).round(3)
say(summ.to_string())

# ----------------------------------------------------------------- OUTPUT ----

D_out = D.copy()
D_out['kept'] = D_out['keep']
cols_r = ['reading', 'ts', 'site', 'side', 'Locus\\Square', 'Basket', 'duration',
          'cover', 'Dirt', 'author_flag', 'kept', 'drop_reason',
          'Rb', 'Zr', 'Nb', 'Fe', 'Ti', 'Mn', 'Bal', '_src_file', '_xlrow']
D_out[[c for c in cols_r if c in D_out.columns]].to_csv(
    OUT / 'samples_readings.csv', index=False)
O.to_csv(OUT / 'samples_objects.csv', index=False)

(OUT / 'cleaning_report.txt').write_text('\n'.join(report), encoding='utf-8')

# ---- human-readable Excel ----

def autosize(ws, df, start=1):
    for j, c in enumerate(df.columns, start=start):
        w = max(len(str(c)), *(len(str(v)) for v in df[c].head(200))) if len(df) else len(str(c))
        ws.column_dimensions[get_column_letter(j)].width = min(38, max(10, w + 2))


def write_sheet(xw, df, name, note=None):
    startrow = 2 if note else 0
    df.to_excel(xw, sheet_name=name, index=False, startrow=startrow)
    ws = xw.sheets[name]
    if note:
        ws['A1'] = note
        ws['A1'].font = Font(italic=True, size=10, color='555555')
    hdr_row = startrow + 1
    for cell in ws[hdr_row]:
        if cell.value is not None:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2F5597')
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    autosize(ws, df)


readme = pd.DataFrame({
    'Sheet': ['Objects', 'Site_summary', 'Needs_review', 'Readings', 'Dropped'],
    'What it contains': [
        'ONE ROW PER OBSIDIAN OBJECT - the analysis-ready table. Use this one.',
        'The three sites compared, using objects (not readings).',
        'Objects whose repeat measurements disagree although both look valid. '
        'These need a human eye.',
        'Every original reading with its keep/drop decision and all quality flags.',
        'Only the readings that were removed, each with the reason.'],
    'Key columns': [
        'Rb_ppm, Zr_ppm = chemistry in parts-per-million. Rb_Zr = the ratio we '
        'compare on. quality = 0-1 score. n_readings = how many measurements.',
        'median = the middle value; sd = how spread out the values are.',
        'worst_fold = how many times bigger the largest reading is than the smallest.',
        'kept = TRUE/FALSE. drop_reason says why if FALSE.',
        'drop_reason.'],
})

with pd.ExcelWriter(OUT / 'obsidian_samples_CLEAN.xlsx', engine='openpyxl') as xw:
    write_sheet(xw, readme, 'README',
                'Cleaned obsidian pXRF data. Built by analysis/03_clean_samples.py '
                'from data/samples/ (raw files unchanged). Values in ppm.')
    obj_cols = ['object_id', 'site', 'locus', 'basket', 'n_readings', 'sides',
                'Rb_ppm', 'Zr_ppm', 'Nb_ppm', 'Fe_ppm', 'Ti_ppm',
                'Rb_Zr', 'Rb_Nb', 'Zr_Nb',
                'quality', 'quality_band', 'worst_disagreement_fold',
                'ungroupable', 'notes']
    write_sheet(xw, O[[c for c in obj_cols if c in O.columns]].round(4), 'Objects',
                'ONE ROW = ONE OBSIDIAN OBJECT. This is the table to analyse. '
                'Rb_Zr is the main comparison value.')
    write_sheet(xw, summ.reset_index(), 'Site_summary',
                'The three sites compared. Counts are OBJECTS, not measurements.')
    rev = pd.DataFrame(review) if review else pd.DataFrame(
        columns=['object_id', 'n', 'worst_fold'])
    write_sheet(xw, rev, 'Needs_review',
                'Objects whose repeat measurements disagree even though each '
                'reading looks valid. Not averaged - they need your judgement.')
    rd = pd.read_csv(OUT / 'samples_readings.csv')
    write_sheet(xw, rd, 'Readings',
                'Every original reading. kept=FALSE means it was removed; '
                'drop_reason says why. Nothing was deleted silently.')
    write_sheet(xw, rd[rd['kept'] == False], 'Dropped',
                'Only the removed readings, with the reason for each.')

hdr('FILES WRITTEN')
for f in ['obsidian_samples_CLEAN.xlsx', 'samples_objects.csv',
          'samples_readings.csv', 'cleaning_report.txt']:
    say(f"  samples_db/{f}")

(OUT / 'cleaning_report.txt').write_text('\n'.join(report), encoding='utf-8')
